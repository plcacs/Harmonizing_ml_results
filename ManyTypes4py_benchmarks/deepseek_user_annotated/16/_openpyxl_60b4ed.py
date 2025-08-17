from __future__ import annotations

import mmap
from typing import (
    TYPE_CHECKING,
    Any,
    Dict,
    Iterable,
    List,
    Optional,
    Tuple,
    Type,
    Union,
    cast,
)

import numpy as np

from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc

from pandas.core.shared_docs import _shared_docs

from pandas.io.excel._base import (
    BaseExcelReader,
    ExcelWriter,
)
from pandas.io.excel._util import (
    combine_kwargs,
    validate_freeze_panes,
)

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.descriptors.serialisable import Serialisable
    from openpyxl.styles import Fill, Font, Color, Side, Border, Alignment, Protection
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell

    from pandas._typing import (
        ExcelWriterIfSheetExists,
        FilePath,
        ReadBuffer,
        Scalar,
        StorageOptions,
        WriteExcelBuffer,
    )


class OpenpyxlWriter(ExcelWriter):
    _engine: str = "openpyxl"
    _supported_extensions: Tuple[str, str] = (".xlsx", ".xlsm")

    def __init__(  # pyright: ignore[reportInconsistentConstructor]
        self,
        path: Union[FilePath, WriteExcelBuffer, ExcelWriter],
        engine: Optional[str] = None,
        date_format: Optional[str] = None,
        datetime_format: Optional[str] = None,
        mode: str = "w",
        storage_options: Optional[StorageOptions] = None,
        if_sheet_exists: Optional[ExcelWriterIfSheetExists] = None,
        engine_kwargs: Optional[Dict[str, Any]] = None,
        **kwargs: Any,
    ) -> None:
        from openpyxl.workbook import Workbook

        engine_kwargs = combine_kwargs(engine_kwargs, kwargs)

        super().__init__(
            path,
            mode=mode,
            storage_options=storage_options,
            if_sheet_exists=if_sheet_exists,
            engine_kwargs=engine_kwargs,
        )

        if "r+" in self._mode:
            from openpyxl import load_workbook

            try:
                self._book: Workbook = load_workbook(self._handles.handle, **engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise
            self._handles.handle.seek(0)
        else:
            try:
                self._book = Workbook(**engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise

            if self.book.worksheets:
                self.book.remove(self.book.worksheets[0])

    @property
    def book(self) -> "Workbook":
        return self._book

    @property
    def sheets(self) -> Dict[str, "Worksheet"]:
        return {name: self.book[name] for name in self.book.sheetnames}

    def _save(self) -> None:
        self.book.save(self._handles.handle)
        if "r+" in self._mode and not isinstance(self._handles.handle, mmap.mmap):
            self._handles.handle.truncate()

    @classmethod
    def _convert_to_style_kwargs(cls, style_dict: Dict[str, Any]) -> Dict[str, "Serialisable"]:
        _style_key_map = {"borders": "border"}

        style_kwargs: Dict[str, "Serialisable"] = {}
        for k, v in style_dict.items():
            k = _style_key_map.get(k, k)
            _conv_to_x = getattr(cls, f"_convert_to_{k}", lambda x: None)
            new_v = _conv_to_x(v)
            if new_v:
                style_kwargs[k] = new_v

        return style_kwargs

    @classmethod
    def _convert_to_color(cls, color_spec: Union[str, Dict[str, Any]]) -> "Color":
        from openpyxl.styles import Color

        if isinstance(color_spec, str):
            return Color(color_spec)
        else:
            return Color(**color_spec)

    @classmethod
    def _convert_to_font(cls, font_dict: Dict[str, Any]) -> "Font":
        from openpyxl.styles import Font

        _font_key_map = {
            "sz": "size",
            "b": "bold",
            "i": "italic",
            "u": "underline",
            "strike": "strikethrough",
            "vertalign": "vertAlign",
        }

        font_kwargs: Dict[str, Any] = {}
        for k, v in font_dict.items():
            k = _font_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            font_kwargs[k] = v

        return Font(**font_kwargs)

    @classmethod
    def _convert_to_stop(cls, stop_seq: Iterable[Any]) -> Iterable["Color"]:
        return map(cls._convert_to_color, stop_seq)

    @classmethod
    def _convert_to_fill(cls, fill_dict: Dict[str, Any]) -> "Fill":
        from openpyxl.styles import GradientFill, PatternFill

        _pattern_fill_key_map = {
            "patternType": "fill_type",
            "patterntype": "fill_type",
            "fgColor": "start_color",
            "fgcolor": "start_color",
            "bgColor": "end_color",
            "bgcolor": "end_color",
        }

        _gradient_fill_key_map = {"fill_type": "type"}

        pfill_kwargs: Dict[str, Any] = {}
        gfill_kwargs: Dict[str, Any] = {}
        for k, v in fill_dict.items():
            pk = _pattern_fill_key_map.get(k)
            gk = _gradient_fill_key_map.get(k)
            if pk in ["start_color", "end_color"]:
                v = cls._convert_to_color(v)
            if gk == "stop":
                v = cls._convert_to_stop(v)
            if pk:
                pfill_kwargs[pk] = v
            elif gk:
                gfill_kwargs[gk] = v
            else:
                pfill_kwargs[k] = v
                gfill_kwargs[k] = v

        try:
            return PatternFill(**pfill_kwargs)
        except TypeError:
            return GradientFill(**gfill_kwargs)

    @classmethod
    def _convert_to_side(cls, side_spec: Union[str, Dict[str, Any]]) -> "Side":
        from openpyxl.styles import Side

        _side_key_map = {"border_style": "style"}

        if isinstance(side_spec, str):
            return Side(style=side_spec)

        side_kwargs: Dict[str, Any] = {}
        for k, v in side_spec.items():
            k = _side_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            side_kwargs[k] = v

        return Side(**side_kwargs)

    @classmethod
    def _convert_to_border(cls, border_dict: Dict[str, Any]) -> "Border":
        from openpyxl.styles import Border

        _border_key_map = {"diagonalup": "diagonalUp", "diagonaldown": "diagonalDown"}

        border_kwargs: Dict[str, Any] = {}
        for k, v in border_dict.items():
            k = _border_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            if k in ["left", "right", "top", "bottom", "diagonal"]:
                v = cls._convert_to_side(v)
            border_kwargs[k] = v

        return Border(**border_kwargs)

    @classmethod
    def _convert_to_alignment(cls, alignment_dict: Dict[str, Any]) -> "Alignment":
        from openpyxl.styles import Alignment

        return Alignment(**alignment_dict)

    @classmethod
    def _convert_to_number_format(cls, number_format_dict: Dict[str, str]) -> str:
        return number_format_dict["format_code"]

    @classmethod
    def _convert_to_protection(cls, protection_dict: Dict[str, Any]) -> "Protection":
        from openpyxl.styles import Protection

        return Protection(**protection_dict)

    def _write_cells(
        self,
        cells: Any,
        sheet_name: Optional[str] = None,
        startrow: int = 0,
        startcol: int = 0,
        freeze_panes: Optional[Tuple[int, int]] = None,
    ) -> None:
        sheet_name = self._get_sheet_name(sheet_name)

        _style_cache: Dict[str, Dict[str, "Serialisable"]] = {}

        if sheet_name in self.sheets and self._if_sheet_exists != "new":
            if "r+" in self._mode:
                if self._if_sheet_exists == "replace":
                    old_wks = self.sheets[sheet_name]
                    target_index = self.book.index(old_wks)
                    del self.book[sheet_name]
                    wks = self.book.create_sheet(sheet_name, target_index)
                elif self._if_sheet_exists == "error":
                    raise ValueError(
                        f"Sheet '{sheet_name}' already exists and "
                        f"if_sheet_exists is set to 'error'."
                    )
                elif self._if_sheet_exists == "overlay":
                    wks = self.sheets[sheet_name]
                else:
                    raise ValueError(
                        f"'{self._if_sheet_exists}' is not valid for if_sheet_exists. "
                        "Valid options are 'error', 'new', 'replace' and 'overlay'."
                    )
            else:
                wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name

        if validate_freeze_panes(freeze_panes):
            freeze_panes = cast(Tuple[int, int], freeze_panes)
            wks.freeze_panes = wks.cell(
                row=freeze_panes[0] + 1, column=freeze_panes[1] + 1
            )

        for cell in cells:
            xcell = wks.cell(
                row=startrow + cell.row + 1, column=startcol + cell.col + 1
            )
            xcell.value, fmt = self._value_with_fmt(cell.val)
            if fmt:
                xcell.number_format = fmt

            style_kwargs: Optional[Dict[str, "Serialisable"]] = None
            if cell.style:
                key = str(cell.style)
                style_kwargs = _style_cache.get(key)
                if style_kwargs is None:
                    style_kwargs = self._convert_to_style_kwargs(cell.style)
                    _style_cache[key] = style_kwargs

            if style_kwargs:
                for k, v in style_kwargs.items():
                    setattr(xcell, k, v)

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.merge_cells(
                    start_row=startrow + cell.row + 1,
                    start_column=startcol + cell.col + 1,
                    end_column=startcol + cell.mergeend + 1,
                    end_row=startrow + cell.mergestart + 1,
                )

                if style_kwargs:
                    first_row = startrow + cell.row + 1
                    last_row = startrow + cell.mergestart + 1
                    first_col = startcol + cell.col + 1
                    last_col = startcol + cell.mergeend + 1

                    for row in range(first_row, last_row + 1):
                        for col in range(first_col, last_col + 1):
                            if row == first_row and col == first_col:
                                continue
                            xcell = wks.cell(column=col, row=row)
                            for k, v in style_kwargs.items():
                                setattr(xcell, k, v)


class OpenpyxlReader(BaseExcelReader["Workbook"]):
    @doc(storage_options=_shared_docs["storage_options"])
    def __init__(
        self,
        filepath_or_buffer: Union[FilePath, ReadBuffer[bytes]],
        storage_options: Optional[StorageOptions] = None,
        engine_kwargs: Optional[Dict[str, Any]] = None,
    ) -> None:
        import_optional_dependency("openpyxl")
        super().__init__(
            filepath_or_buffer,
            storage_options=storage_options,
            engine_kwargs=engine_kwargs,
        )

    @property
    def _workbook_class(self) -> Type["Workbook"]:
        from openpyxl import Workbook
        return Workbook

    def load_workbook(
        self, 
        filepath_or_buffer: Union[FilePath, ReadBuffer[bytes]], 
        engine_kwargs: Dict[str, Any]
    ) -> "Workbook":
        from openpyxl import load_workbook

        default_kwargs = {"read_only": True, "data_only": True, "keep_links": False}
        return load_workbook(
            filepath_or_buffer,
            **(default_kwargs | engine_kwargs),
        )

    @property
    def sheet_names(self) -> List[str]:
        return [sheet.title for sheet in self.book.worksheets]

    def get_sheet_by_name(self, name: str) -> "Worksheet":
        self.raise_if_bad_sheet_by_name(name)
        return self.book[name]

    def get_sheet_by_index(self, index: int) -> "Worksheet":
        self.raise_if_bad_sheet_by_index(index)
        return self.book.worksheets[index]

    def _convert_cell(self, cell: "Cell") -> Scalar:
        from openpyxl.cell.cell import TYPE_ERROR, TYPE_NUMERIC

        if cell.value is None:
            return ""
        elif cell.data_type == TYPE_ERROR:
            return np.nan
        elif cell.data_type == TYPE_NUMERIC:
            val = int(cell.value)
            if val == cell.value:
                return val
            return float(cell.value)
        return cell.value

    def get_sheet_data(
        self, 
        sheet: "Worksheet", 
        file_rows_needed: Optional[int] = None
    ) -> List[List[Scalar]]:
        if self.book.read_only:
            sheet.reset_dimensions()

        data: List[List[Scalar]] = []
        last_row_with_data = -1
        for row_number, row in enumerate(sheet.rows):
            converted_row = [self._convert_cell(cell) for cell in row]
            while converted_row and converted_row[-1] == "":
                converted_row.pop()
            if converted_row:
                last_row_with_data = row_number
            data.append(converted_row)
            if file_rows_needed is not None and len(data) >= file_rows_needed:
                break

        data = data[: last_row_with_data + 1]

        if len(data) > 0:
            max_width = max(len(data_row) for data_row in data)
            if min(len(data_row) for data_row in data) < max_width:
                empty_cell: List[Scalar] = [""]
                data = [
                    data_row + (max_width - len(data_row)) * empty_cell
                    for data_row in data
                ]

        return data
