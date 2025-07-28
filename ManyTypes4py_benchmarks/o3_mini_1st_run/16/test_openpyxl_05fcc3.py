#!/usr/bin/env python3
from __future__ import annotations
import contextlib
from pathlib import Path
import re
import uuid
import numpy as np
import pytest
import pandas as pd
from pandas import DataFrame
import pandas._testing as tm
from pandas.io.excel import ExcelWriter, _OpenpyxlWriter
from pandas.io.excel._openpyxl import OpenpyxlReader
from typing import Any, Dict, List, Optional

openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture
def ext() -> str:
    return ".xlsx"


@pytest.fixture
def tmp_excel(ext: str, tmp_path: Path) -> str:
    tmp: Path = tmp_path / f"{uuid.uuid4()}{ext}"
    tmp.touch()
    return str(tmp)


def test_to_excel_styleconverter() -> None:
    from openpyxl import styles

    hstyle: Dict[str, Any] = {
        "font": {"color": "00FF0000", "bold": True},
        "borders": {"top": "thin", "right": "thin", "bottom": "thin", "left": "thin"},
        "alignment": {"horizontal": "center", "vertical": "top"},
        "fill": {"patternType": "solid", "fgColor": {"rgb": "006666FF", "tint": 0.3}},
        "number_format": {"format_code": "0.00"},
        "protection": {"locked": True, "hidden": False},
    }
    font_color: styles.Color = styles.Color("00FF0000")
    font: styles.Font = styles.Font(bold=True, color=font_color)
    side: styles.Side = styles.Side(style=styles.borders.BORDER_THIN)
    border: styles.Border = styles.Border(top=side, right=side, bottom=side, left=side)
    alignment: styles.Alignment = styles.Alignment(horizontal="center", vertical="top")
    fill_color: styles.Color = styles.Color(rgb="006666FF", tint=0.3)
    fill: styles.PatternFill = styles.PatternFill(patternType="solid", fgColor=fill_color)
    number_format: str = "0.00"
    protection: styles.Protection = styles.Protection(locked=True, hidden=False)
    kw: Dict[str, Any] = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
    assert kw["font"] == font
    assert kw["border"] == border
    assert kw["alignment"] == alignment
    assert kw["fill"] == fill
    assert kw["number_format"] == number_format
    assert kw["protection"] == protection


def test_write_cells_merge_styled(tmp_excel: str) -> None:
    from pandas.io.formats.excel import ExcelCell

    sheet_name: str = "merge_styled"
    sty_b1: Dict[str, Any] = {"font": {"color": "00FF0000"}}
    sty_a2: Dict[str, Any] = {"font": {"color": "0000FF00"}}
    initial_cells: List[ExcelCell] = [
        ExcelCell(col=1, row=0, val=42, style=sty_b1),
        ExcelCell(col=0, row=1, val=99, style=sty_a2),
    ]
    sty_merged: Dict[str, Any] = {"font": {"color": "000000FF", "bold": True}}
    sty_kwargs: Dict[str, Any] = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
    openpyxl_sty_merged: Any = sty_kwargs["font"]
    merge_cells: List[ExcelCell] = [
        ExcelCell(col=0, row=0, val="pandas", mergestart=1, mergeend=1, style=sty_merged)
    ]
    with _OpenpyxlWriter(tmp_excel) as writer:
        writer._write_cells(initial_cells, sheet_name=sheet_name)
        writer._write_cells(merge_cells, sheet_name=sheet_name)
        wks = writer.sheets[sheet_name]
    xcell_b1 = wks["B1"]
    xcell_a2 = wks["A2"]
    assert xcell_b1.font == openpyxl_sty_merged
    assert xcell_a2.font == openpyxl_sty_merged


@pytest.mark.parametrize("iso_dates", [True, False])
def test_engine_kwargs_write(tmp_excel: str, iso_dates: bool) -> None:
    engine_kwargs: Dict[str, Any] = {"iso_dates": iso_dates}
    with ExcelWriter(tmp_excel, engine="openpyxl", engine_kwargs=engine_kwargs) as writer:
        assert writer.book.iso_dates == iso_dates
        DataFrame().to_excel(writer)


def test_engine_kwargs_append_invalid(tmp_excel: str) -> None:
    DataFrame(["hello", "world"]).to_excel(tmp_excel)
    with pytest.raises(
        TypeError,
        match=re.escape("load_workbook() got an unexpected keyword argument 'apple_banana'"),
    ):
        with ExcelWriter(tmp_excel, engine="openpyxl", mode="a", engine_kwargs={"apple_banana": "fruit"}) as writer:
            DataFrame(["good"]).to_excel(writer, sheet_name="Sheet2")


@pytest.mark.parametrize("data_only, expected", [(True, 0), (False, "=1+1")])
def test_engine_kwargs_append_data_only(tmp_excel: str, data_only: bool, expected: Any) -> None:
    DataFrame(["=1+1"]).to_excel(tmp_excel)
    with ExcelWriter(tmp_excel, engine="openpyxl", mode="a", engine_kwargs={"data_only": data_only}) as writer:
        assert writer.sheets["Sheet1"]["B2"].value == expected
        DataFrame().to_excel(writer, sheet_name="Sheet2")
    assert pd.read_excel(
        tmp_excel, sheet_name="Sheet1", engine="openpyxl", engine_kwargs={"data_only": data_only}
    ).iloc[0, 1] == expected


@pytest.mark.parametrize("kwarg_name", ["read_only", "data_only"])
@pytest.mark.parametrize("kwarg_value", [True, False])
def test_engine_kwargs_append_reader(
    datapath: Any, ext: str, kwarg_name: str, kwarg_value: bool
) -> None:
    filename: str = datapath("io", "data", "excel", "test1" + ext)
    with contextlib.closing(OpenpyxlReader(filename, engine_kwargs={kwarg_name: kwarg_value})) as reader:
        assert getattr(reader.book, kwarg_name) == kwarg_value


@pytest.mark.parametrize("mode,expected", [("w", ["baz"]), ("a", ["foo", "bar", "baz"])])
def test_write_append_mode(tmp_excel: str, mode: str, expected: List[str]) -> None:
    df: DataFrame = DataFrame([1], columns=["baz"])
    wb = openpyxl.Workbook()
    wb.worksheets[0].title = "foo"
    wb.worksheets[0]["A1"].value = "foo"
    wb.create_sheet("bar")
    wb.worksheets[1]["A1"].value = "bar"
    wb.save(tmp_excel)
    with ExcelWriter(tmp_excel, engine="openpyxl", mode=mode) as writer:
        df.to_excel(writer, sheet_name="baz", index=False)
    with contextlib.closing(openpyxl.load_workbook(tmp_excel)) as wb2:
        result: List[str] = [sheet.title for sheet in wb2.worksheets]
        assert result == expected
        for index, cell_value in enumerate(expected):
            assert wb2.worksheets[index]["A1"].value == cell_value


@pytest.mark.parametrize(
    "if_sheet_exists,num_sheets,expected",
    [
        ("new", 2, ["apple", "banana"]),
        ("replace", 1, ["pear"]),
        ("overlay", 1, ["pear", "banana"]),
    ],
)
def test_if_sheet_exists_append_modes(
    tmp_excel: str, if_sheet_exists: str, num_sheets: int, expected: List[str]
) -> None:
    df1: DataFrame = DataFrame({"fruit": ["apple", "banana"]})
    df2: DataFrame = DataFrame({"fruit": ["pear"]})
    df1.to_excel(tmp_excel, engine="openpyxl", sheet_name="foo", index=False)
    with ExcelWriter(tmp_excel, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists) as writer:
        df2.to_excel(writer, sheet_name="foo", index=False)
    with contextlib.closing(openpyxl.load_workbook(tmp_excel)) as wb:
        assert len(wb.sheetnames) == num_sheets
        assert wb.sheetnames[0] == "foo"
        result: DataFrame = pd.read_excel(wb, "foo", engine="openpyxl")
        assert list(result["fruit"]) == expected
        if len(wb.sheetnames) == 2:
            result2: DataFrame = pd.read_excel(wb, wb.sheetnames[1], engine="openpyxl")
            tm.assert_frame_equal(result2, df2)


@pytest.mark.parametrize(
    "startrow, startcol, greeting, goodbye",
    [
        (0, 0, ["poop", "world"], ["goodbye", "people"]),
        (0, 1, ["hello", "world"], ["poop", "people"]),
        (1, 0, ["hello", "poop"], ["goodbye", "people"]),
        (1, 1, ["hello", "world"], ["goodbye", "poop"]),
    ],
)
def test_append_overlay_startrow_startcol(
    tmp_excel: str, startrow: int, startcol: int, greeting: List[str], goodbye: List[str]
) -> None:
    df1: DataFrame = DataFrame({"greeting": ["hello", "world"], "goodbye": ["goodbye", "people"]})
    df2: DataFrame = DataFrame(["poop"])
    df1.to_excel(tmp_excel, engine="openpyxl", sheet_name="poo", index=False)
    with ExcelWriter(tmp_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df2.to_excel(writer, index=False, header=False, startrow=startrow + 1, startcol=startcol, sheet_name="poo")
    result: DataFrame = pd.read_excel(tmp_excel, sheet_name="poo", engine="openpyxl")
    expected: DataFrame = DataFrame({"greeting": greeting, "goodbye": goodbye})
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "if_sheet_exists,msg",
    [
        ("invalid", "'invalid' is not valid for if_sheet_exists. Valid options are 'error', 'new', 'replace' and 'overlay'."),
        ("error", "Sheet 'foo' already exists and if_sheet_exists is set to 'error'."),
        (None, "Sheet 'foo' already exists and if_sheet_exists is set to 'error'."),
    ],
)
def test_if_sheet_exists_raises(tmp_excel: str, if_sheet_exists: Optional[str], msg: str) -> None:
    df: DataFrame = DataFrame({"fruit": ["pear"]})
    df.to_excel(tmp_excel, sheet_name="foo", engine="openpyxl")
    with pytest.raises(ValueError, match=re.escape(msg)):
        with ExcelWriter(tmp_excel, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name="foo")


def test_to_excel_with_openpyxl_engine(tmp_excel: str) -> None:
    df1: DataFrame = DataFrame({"A": np.linspace(1, 10, 10)})
    df2: DataFrame = DataFrame({"B": np.linspace(1, 20, 10)})
    df: DataFrame = pd.concat([df1, df2], axis=1)
    styled = df.style.map(lambda val: f"color: {('red' if val < 0 else 'black')}").highlight_max()
    styled.to_excel(tmp_excel, engine="openpyxl")


@pytest.mark.parametrize("read_only", [True, False])
def test_read_workbook(datapath: Any, ext: str, read_only: bool) -> None:
    filename: str = datapath("io", "data", "excel", "test1" + ext)
    with contextlib.closing(openpyxl.load_workbook(filename, read_only=read_only)) as wb:
        result: DataFrame = pd.read_excel(wb, engine="openpyxl")
    expected: DataFrame = pd.read_excel(filename)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "header, expected_data",
    [
        (
            0,
            {
                "Title": [np.nan, "A", 1, 2, 3],
                "Unnamed: 1": [np.nan, "B", 4, 5, 6],
                "Unnamed: 2": [np.nan, "C", 7, 8, 9],
            },
        ),
        (
            2,
            {"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]},
        ),
    ],
)
@pytest.mark.parametrize("filename", ["dimension_missing", "dimension_small", "dimension_large"])
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_bad_dimension(
    datapath: Any, ext: str, header: int, expected_data: Dict[str, List[Any]], filename: str, read_only: Optional[bool]
) -> None:
    path: str = datapath("io", "data", "excel", f"{filename}{ext}")
    if read_only is None:
        result: DataFrame = pd.read_excel(path, header=header)
    else:
        with contextlib.closing(openpyxl.load_workbook(path, read_only=read_only)) as wb:
            result = pd.read_excel(wb, engine="openpyxl", header=header)
    expected: DataFrame = DataFrame(expected_data)
    tm.assert_frame_equal(result, expected)


def test_append_mode_file(tmp_excel: str) -> None:
    df: DataFrame = DataFrame()
    df.to_excel(tmp_excel, engine="openpyxl")
    with ExcelWriter(tmp_excel, mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
        df.to_excel(writer)
    data: bytes = Path(tmp_excel).read_bytes()
    first: int = data.find(b"docProps/app.xml")
    second: int = data.find(b"docProps/app.xml", first + 1)
    third: int = data.find(b"docProps/app.xml", second + 1)
    assert second != -1 and third == -1


@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_empty_trailing_rows(datapath: Any, ext: str, read_only: Optional[bool]) -> None:
    path: str = datapath("io", "data", "excel", f"empty_trailing_rows{ext}")
    if read_only is None:
        result: DataFrame = pd.read_excel(path)
    else:
        with contextlib.closing(openpyxl.load_workbook(path, read_only=read_only)) as wb:
            result = pd.read_excel(wb, engine="openpyxl")
    expected: DataFrame = DataFrame(
        {
            "Title": [np.nan, "A", 1, 2, 3],
            "Unnamed: 1": [np.nan, "B", 4, 5, 6],
            "Unnamed: 2": [np.nan, "C", 7, 8, 9],
        }
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_empty_with_blank_row(datapath: Any, ext: str, read_only: Optional[bool]) -> None:
    path: str = datapath("io", "data", "excel", f"empty_with_blank_row{ext}")
    if read_only is None:
        result: DataFrame = pd.read_excel(path)
    else:
        with contextlib.closing(openpyxl.load_workbook(path, read_only=read_only)) as wb:
            result = pd.read_excel(wb, engine="openpyxl")
    expected: DataFrame = DataFrame()
    tm.assert_frame_equal(result, expected)


def test_book_and_sheets_consistent(tmp_excel: str) -> None:
    with ExcelWriter(tmp_excel, engine="openpyxl") as writer:
        assert writer.sheets == {}
        sheet = writer.book.create_sheet("test_name", 0)
        assert writer.sheets == {"test_name": sheet}


def test_ints_spelled_with_decimals(datapath: Any, ext: str) -> None:
    path: str = datapath("io", "data", "excel", f"ints_spelled_with_decimals{ext}")
    result: DataFrame = pd.read_excel(path)
    expected: DataFrame = DataFrame(range(2, 12), columns=[1])
    tm.assert_frame_equal(result, expected)


def test_read_multiindex_header_no_index_names(datapath: Any, ext: str) -> None:
    path: str = datapath("io", "data", "excel", f"multiindex_no_index_names{ext}")
    result: DataFrame = pd.read_excel(path, index_col=[0, 1, 2], header=[0, 1, 2])
    expected: DataFrame = DataFrame(
        [
            [np.nan, "x", "x", "x"],
            ["x", np.nan, np.nan, np.nan],
        ],
        columns=pd.MultiIndex.from_tuples(
            [("X", "Y", "A1"), ("X", "Y", "A2"), ("XX", "YY", "B1"), ("XX", "YY", "B2")]
        ),
        index=pd.MultiIndex.from_tuples([("A", "AA", "AAA"), ("A", "BB", "BBB")]),
    )
    tm.assert_frame_equal(result, expected)