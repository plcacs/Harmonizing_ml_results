from __future__ import annotations
from typing import TYPE_CHECKING, Any, cast, Dict, List, Tuple
import mmap
import numpy as np
from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc
from pandas.core.shared_docs import _shared_docs
from pandas.io.excel._base import BaseExcelReader, ExcelWriter
from pandas.io.excel._util import combine_kwargs, validate_freeze_panes
if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.descriptors.serialisable import Serialisable
    from openpyxl.styles import Fill
    from pandas._typing import ExcelWriterIfSheetExists, FilePath, ReadBuffer, Scalar, StorageOptions, WriteExcelBuffer

class OpenpyxlWriter(ExcelWriter):
    _engine: str = 'openpyxl'
    _supported_extensions: Tuple[str, str] = ('.xlsx', '.xlsm')

    def __init__(self, path: FilePath, engine: Any = None, date_format: Any = None, datetime_format: Any = None, mode: str = 'w', storage_options: StorageOptions = None, if_sheet_exists: Any = None, engine_kwargs: Dict[str, Any] = None, **kwargs: Any) -> None:
        from openpyxl.workbook import Workbook
        engine_kwargs = combine_kwargs(engine_kwargs, kwargs)
        super().__init__(path, mode=mode, storage_options=storage_options, if_sheet_exists=if_sheet_exists, engine_kwargs=engine_kwargs)
        if 'r+' in self._mode:
            from openpyxl import load_workbook
            try:
                self._book = load_workbook(self._handles.handle, **engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise
            self._handles.handle.seek(0)
        else:
            try:
                self._book = Workbook(**engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise
            if self.book.worksheets:
                self.book.remove(self.book.worksheets[0])

    @property
    def book(self) -> Workbook:
        """
        Book instance of class openpyxl.workbook.Workbook.

        This attribute can be used to access engine-specific features.
        """
        return self._book

    @property
    def sheets(self) -> Dict[str, Any]:
        """Mapping of sheet names to sheet objects."""
        result = {name: self.book[name] for name in self.book.sheetnames}
        return result

    def _save(self) -> None:
        """
        Save workbook to disk.
        """
        self.book.save(self._handles.handle)
        if 'r+' in self._mode and (not isinstance(self._handles.handle, mmap.mmap)):
            self._handles.handle.truncate()

    @classmethod
    def _convert_to_style_kwargs(cls, style_dict: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert a style_dict to a set of kwargs suitable for initializing
        or updating-on-copy an openpyxl v2 style object.

        Parameters
        ----------
        style_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'font'
                'fill'
                'border' ('borders')
                'alignment'
                'number_format'
                'protection'

        Returns
        -------
        style_kwargs : dict
            A dict with the same, normalized keys as ``style_dict`` but each
            value has been replaced with a native openpyxl style object of the
            appropriate class.
        """
        _style_key_map = {'borders': 'border'}
        style_kwargs = {}
        for k, v in style_dict.items():
            k = _style_key_map.get(k, k)
            _conv_to_x = getattr(cls, f'_convert_to_{k}', lambda x: None)
            new_v = _conv_to_x(v)
            if new_v:
                style_kwargs[k] = new_v
        return style_kwargs

    @classmethod
    def _convert_to_color(cls, color_spec: Any) -> Any:
        """
        Convert ``color_spec`` to an openpyxl v2 Color object.

        Parameters
        ----------
        color_spec : str, dict
            A 32-bit ARGB hex string, or a dict with zero or more of the
            following keys.
                'rgb'
                'indexed'
                'auto'
                'theme'
                'tint'
                'index'
                'type'

        Returns
        -------
        color : openpyxl.styles.Color
        """
        from openpyxl.styles import Color
        if isinstance(color_spec, str):
            return Color(color_spec)
        else:
            return Color(**color_spec)

    # Remaining methods have similar type annotations
