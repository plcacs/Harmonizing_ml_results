import contextlib
from pathlib import Path
import re
import uuid
import numpy as np
import pytest
import pandas as pd
from pandas import DataFrame
import pandas._testing as tm
from pandas.io.excel import ExcelWriter, _OpenpyxlWriter
from pandas.io.excel._openpyxl import OpenpyxlReader
openpyxl = pytest.importorskip('openpyxl')

@pytest.fixture
def ext() -> str:
    return '.xlsx'

@pytest.fixture
def tmp_excel(ext: str, tmp_path: Path) -> str:
    tmp = tmp_path / f'{uuid.uuid4()}{ext}'
    tmp.touch()
    return str(tmp)

def test_to_excel_styleconverter():
    from openpyxl import styles
    hstyle = {'font': {'color': '00FF0000', 'bold': True}, 'borders': {'top': 'thin', 'right': 'thin', 'bottom': 'thin', 'left': 'thin'}, 'alignment': {'horizontal': 'center', 'vertical': 'top'}, 'fill': {'patternType': 'solid', 'fgColor': {'rgb': '006666FF', 'tint': 0.3}}, 'number_format': {'format_code': '0.00'}, 'protection': {'locked': True, 'hidden': False}}
    font_color = styles.Color('00FF0000')
    font = styles.Font(bold=True, color=font_color)
    side = styles.Side(style=styles.borders.BORDER_THIN)
    border = styles.Border(top=side, right=side, bottom=side, left=side)
    alignment = styles.Alignment(horizontal='center', vertical='top')
    fill_color = styles.Color(rgb='006666FF', tint=0.3)
    fill = styles.PatternFill(patternType='solid', fgColor=fill_color)
    number_format = '0.00'
    protection = styles.Protection(locked=True, hidden=False)
    kw = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
    assert kw['font'] == font
    assert kw['border'] == border
    assert kw['alignment'] == alignment
    assert kw['fill'] == fill
    assert kw['number_format'] == number_format
    assert kw['protection'] == protection

def test_write_cells_merge_styled(tmp_excel: str):
    from pandas.io.formats.excel import ExcelCell
    sheet_name = 'merge_styled'
    sty_b1 = {'font': {'color': '00FF0000'}}
    sty_a2 = {'font': {'color': '0000FF00'}}
    initial_cells = [ExcelCell(col=1, row=0, val=42, style=sty_b1), ExcelCell(col=0, row=1, val=99, style=sty_a2)]
    sty_merged = {'font': {'color': '000000FF', 'bold': True}}
    sty_kwargs = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
    openpyxl_sty_merged = sty_kwargs['font']
    merge_cells = [ExcelCell(col=0, row=0, val='pandas', mergestart=1, mergeend=1, style=sty_merged)]
    with _OpenpyxlWriter(tmp_excel) as writer:
        writer._write_cells(initial_cells, sheet_name=sheet_name)
        writer._write_cells(merge_cells, sheet_name=sheet_name)
        wks = writer.sheets[sheet_name]
    xcell_b1 = wks['B1']
    xcell_a2 = wks['A2']
    assert xcell_b1.font == openpyxl_sty_merged
    assert xcell_a2.font == openpyxl_sty_merged

@pytest.mark.parametrize('iso_dates', [True, False])
def test_engine_kwargs_write(tmp_excel: str, iso_dates: bool):
    engine_kwargs = {'iso_dates': iso_dates}
    with ExcelWriter(tmp_excel, engine='openpyxl', engine_kwargs=engine_kwargs) as writer:
        assert writer.book.iso_dates == iso_dates
        DataFrame().to_excel(writer)
